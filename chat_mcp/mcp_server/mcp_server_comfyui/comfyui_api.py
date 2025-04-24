# -*- coding: utf-8 -*-
import asyncio
import logging
import os
import random
import websocket
import json
import requests
import time
import threading
from pathlib import Path
from typing import Optional, Tuple, Dict, Any, Callable
from openpyxl.reader.excel import load_workbook
from chat_mcp.utils.get_logger import get_logger
from config.config import COMFYUI_PATH, COMFYUI_MODEL_PATH

logger = logging.getLogger("ComfyuiAPI")

class ComfyuiAPI:
    """
    Comfyui图像生成客户端类
    """
    _instance = None
    _initialized = False

    def __new__(cls, *args, **kwargs):
        """实现单例模式"""
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self,server_url: str = "127.0.0.1:8188", client_id: str = "123456"):
        """初始化图像客户端"""
        if not ComfyuiAPI._initialized:
            self.server_url = server_url
            self.client_id = client_id
            self.ws = None
            self.is_connected = False
            self.connection_event = threading.Event()
            self.loading = None
            self.ws_thread = None
            self.current_prompt_id = None  # 跟踪当前任务ID
            self.is_generating = False     # 跟踪生成状态
            self._setup_websocket()
            if not self.ensure_connection(timeout=10):
                print("警告：WebSocket连接未能在初始化时建立")

            ComfyuiAPI._initialized = True

    def reinit_connection(self) -> bool:
        """重新初始化连接"""
        try:
            self.close()
            self.ws = None
            self.is_connected = False
            self.connection_event.clear()
            self.loading = None
            self.current_prompt_id = None
            self.is_generating = False

            self._setup_websocket()
            return self.ensure_connection(timeout=10)
        except Exception as e:
            logger.error(f"重新初始化连接时出错: {e}")
            return False

    def _setup_websocket(self) -> None:
        """设置并启动WebSocket连接"""
        ws_url = f"ws://{self.server_url}/ws?clientId={self.client_id}"
        self.ws = websocket.WebSocketApp(
            ws_url,
            on_message=self._on_message,
            on_error=self._on_error,
            on_close=self._on_close,
            on_open=self._on_open
        )

        self.ws_thread = threading.Thread(target=self._run_websocket)
        self.ws_thread.daemon = True
        self.ws_thread.start()

    def _run_websocket(self) -> None:
        """在独立线程中运行WebSocket"""
        try:
            self.ws.run_forever()
        except Exception as e:
            logger.error(f"WebSocket运行错误: {e}")
            self.is_connected = False
            self.connection_event.clear()

    def _on_open(self, ws) -> None:
        """WebSocket连接建立时的回调"""
        self.is_connected = True
        self.connection_event.set()

    def _on_error(self, ws, error) -> None:
        """WebSocket错误处理回调"""
        logger.info(f"WebSocket错误: {error}")
        self.is_connected = False
        self.connection_event.clear()
        if self.loading:
            self.loading.stop()
            self.loading = None

    def _on_close(self, ws, close_status_code, close_msg) -> None:
        """WebSocket关闭时的回调"""
        logger.info(f"WebSocket连接已关闭（状态码：{close_status_code}，消息：{close_msg}）")
        self.is_connected = False
        self.connection_event.clear()
        if self.loading:
            self.loading.stop()
            self.loading = None

    def _on_message(self, ws, message) -> None:
        """处理WebSocket消息"""
        try:
            if isinstance(message, bytes):
                return

            data = json.loads(message)
            msg_type = data.get("type")

            if msg_type == "executing":
                execution_data = data.get("data", {})
                if execution_data.get("prompt_id"):
                    prompt_id = execution_data["prompt_id"]
                    self.current_prompt_id = prompt_id
                    self.is_generating = True
                    self._handle_generation_complete(prompt_id)
            elif msg_type == "progress":
                # 可以在这里处理进度信息
                pass
            elif msg_type == "executed":
                self.is_generating = False
                self.current_prompt_id = None

        except Exception as e:
            logger.error(f"处理WebSocket消息时出错: {e}")

    def _handle_generation_complete(self, prompt_id: str, check_stop: Callable[[], bool] = None) -> None:
        """处理生成完成事件"""
        try:
            if check_stop and check_stop():
                return
            # 获取生成历史
            response = requests.get(
                f"http://{self.server_url}/history/{prompt_id}",
                timeout=10
            )
            if not response.ok:
                return

            if check_stop and check_stop():
                return

            history_data = response.json()
            if prompt_id not in history_data:
                return

            # 获取所有输出节点的图像
            all_images = []
            for node_output in history_data[prompt_id].get("outputs", {}).values():
                if "images" in node_output:
                    all_images.extend(node_output["images"])

            if not all_images:
                return

            # 设置保存路径
            if hasattr(self, '_save_path'):
                save_path = self._save_path
            else:
                save_path = Path("output")
            save_path.mkdir(parents=True, exist_ok=True)

            # 保存所有图片
            saved_paths = []
            success = False

            for idx, image_data in enumerate(all_images):
                image_filename = image_data["filename"]
                image_url = f"http://{self.server_url}/view?filename={image_filename}&type=output"

                img_response = requests.get(image_url, stream=True, timeout=30)
                if not img_response.ok:
                    continue

                # 构建保存文件名
                if hasattr(self, '_image_filename'):
                    base_name = Path(self._image_filename).stem
                    extension = Path(self._image_filename).suffix
                    save_filename = f"{base_name}_{idx}{extension}" if idx > 0 else self._image_filename
                else:
                    timestamp = time.strftime("%Y%m%d_%H%M%S")
                    save_filename = f"image_{timestamp}_{idx}.png"

                full_path = save_path / save_filename

                try:
                    with open(full_path, 'wb') as f:
                        for chunk in img_response.iter_content(chunk_size=8192):
                            if chunk:
                                f.write(chunk)
                    saved_paths.append(str(full_path))
                    success = True
                except Exception as e:
                    logger.error(f"保存图片失败 {save_filename}: {e}")

            # 设置图片路径到实例变量中
            self._generated_paths = saved_paths if saved_paths else None

            # 回调处理
            if hasattr(self, '_save_callback'):
                self._save_callback(success)

        except Exception as e:
            logger.error(f"处理生成完成事件失败: {e}")
            if hasattr(self, '_save_callback'):
                self._save_callback(False)

    def stop_current_task(self) -> bool:
        """停止当前正在运行的任务"""
        try:
            if not self.is_generating:
                logger.info("没有正在运行的任务")
                return False

            # 发送停止请求到ComfyUI服务器
            response = requests.post(
                f"http://{self.server_url}/interrupt",
                timeout=10
            )

            if response.ok:
                logger.info("成功停止当前任务")
                self.is_generating = False
                self.current_prompt_id = None
                return True
            else:
                logger.error(f"停止任务失败: {response.status_code}")
                return False

        except Exception as e:
            logger.error(f"停止任务时出错: {e}")
            return False

    def get_generation_status(self) -> bool:
        """获取当前生成状态"""
        return self.is_generating

    def _load_workflow(self, chinese_name: str) -> Tuple[Optional[Dict], Optional[str]]:
        """加载工作流配置文件"""
        try:
            # 直接使用openpyxl读取Excel，而不是pandas
            from openpyxl import load_workbook

            wb = load_workbook("config/workflow_json.xlsx", read_only=True, data_only=True)
            ws = wb.active

            # 找到列索引
            header_row = next(ws.rows)
            col_indices = {}
            for i, cell in enumerate(header_row):
                if cell.value:
                    col_indices[cell.value] = i

            if 'chinese_name' not in col_indices:
                logger.error("工作流配置文件缺少'chinese_name'列")
                return None, None

            # 查找匹配行
            target_row = None
            for row in list(ws.rows)[1:]:  # 跳过表头
                if row[col_indices['chinese_name']].value == chinese_name:
                    target_row = row
                    break

            # 如果没找到，使用第一行作为默认
            if not target_row and len(list(ws.rows)) > 1:
                logger.info(f"未找到工作流配置：{chinese_name}，使用默认配置")
                target_row = list(ws.rows)[1]  # 第一个数据行

            if not target_row:
                logger.error("无法读取工作流配置")
                return None, None

            json_name = target_row[col_indices['json_name']].value
            checkpoint = target_row[col_indices['checkpoint']].value

            file_path = Path("config/json_file") / f"{json_name}.json"
            if not file_path.exists():
                logger.info(f"找不到配置文件: {file_path}")
                return None, None

            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    prompt_data = json.load(f)
                    workflow_data = {
                        "client_id": self.client_id,
                        "prompt": prompt_data
                    }
                    return workflow_data, checkpoint
            except json.JSONDecodeError as e:
                logger.error(f"JSON解析错误: {e}, 文件: {file_path}")
                return None, None

        except Exception as e:
            logger.error(f"加载工作流配置文件失败: {e}")
            return None, None

    def _update_workflow_params(self, data: Any, params: Dict) -> None:
        """更新工作流参数"""
        try:
            def recursive_update(data_item: Any) -> None:
                if isinstance(data_item, dict):
                    for key, value in data_item.items():
                        if isinstance(value, str) and value.endswith('_data_json_workflow'):
                            if key == 'seed':
                                data_item[key] = params.get(value)
                            else:
                                data_item[key] = params.get(value)
                        elif isinstance(value, (dict, list)):
                            recursive_update(value)
                elif isinstance(data_item, list):
                    for item in data_item:
                        if isinstance(item, (dict, list)):
                            recursive_update(item)

            recursive_update(data)

        except Exception as e:
            logger.info(f"参数更新错误: {str(e)}")

    def get_type_mode(self, excel_path: str) -> list:
        """获取可用的工作流类型"""
        try:
            wb = load_workbook(excel_path, read_only=True)
            ws = wb.active
            header_row = next(ws.rows)
            chinese_name_col = None
            for i, cell in enumerate(header_row):
                if cell.value == 'chinese_name':
                    chinese_name_col = i
                    break

            if chinese_name_col is None:
                logger.error("错误：找不到 'chinese_name' 列")
                return []

            chinese_names = []
            for row in ws.iter_rows(min_row=2):
                cell_value = row[chinese_name_col].value
                if cell_value:
                    chinese_names.append(cell_value)

            return chinese_names

        except Exception as e:
            logger.error(f"获取工作流类型失败: {e}")
            return []

    def wait_for_connection(self, timeout: int = 10) -> bool:
        """等待WebSocket连接建立"""
        return self.connection_event.wait(timeout)

    def ensure_connection(self, timeout: int = 10) -> bool:
        """确保WebSocket连接已建立"""
        if self.is_connected:
            return True

        try:
            response = requests.get(f"http://{self.server_url}/", timeout=5)
            if not response.ok:
                logger.error("错误：ComfyUI服务器未响应")
                return False
        except requests.exceptions.RequestException as e:
            logger.error(f"错误：无法连接到ComfyUI服务器：{e}")
            return False

        if not self.wait_for_connection(timeout):
            logger.error(f"错误：WebSocket连接超时（{timeout}秒）")
            return False

        return True

    def get_models(self, models_dir: str = os.path.join(COMFYUI_PATH, COMFYUI_MODEL_PATH)) -> list:
        """获取可用的模型列表"""
        model_files = []
        if os.path.exists(models_dir):
            for file in os.listdir(models_dir):
                if os.path.isfile(os.path.join(models_dir, file)) and file.endswith('.safetensors'):
                    model_files.append(file)
        return model_files

    def close(self) -> None:
        """关闭连接并清理资源"""
        try:
            if self.ws:
                self.ws.close()
            if self.loading:
                self.loading.stop()
            self.is_connected = False
            self.is_generating = False
            self.current_prompt_id = None
            self.connection_event.clear()
            logger.info("已关闭所有连接并清理资源")
        except Exception as e:
            logger.error(f"关闭连接时出错: {e}")

    async def run(self,
            task_mode: str = "基础文生图",
            model_name: str = None,
            prompt_text: str = None,
            image_name: Optional[str] = None,
            negative_prompt_text: str = None,
            width: int = 512,
            height: int = 768,
            steps: int = 20,
            batch_count: int = 1,
            batch_size: int = 4,
            sampling_method: str = "euler",
            cfg_scale: float = 8.0,
            output_dir: Optional[str] = "output",
            custom_filename: Optional[str] = None,
            timeout: int = 60,
            check_stop: Callable[[], bool] = None,
            ) -> Optional[list]:
        """运行图像生成任务，返回所有生成图片的路径列表

        Args:
            task_mode (str): 任务模式，默认为"基础文生图"
            model_name (str): 模型名称
            prompt_text (str): 提示文本
            image_name (str, optional): 图片名称
            negative_prompt_text (str, optional): 负面提示文本
            width (int): 图片宽度
            height (int): 图片高度
            steps (int): 生成步数
            batch_count (int): 批次数量
            batch_size (int): 每批次大小
            sampling_method (str): 采样方法
            cfg_scale (float): CFG比例
            output_dir (str, optional): 输出目录
            custom_filename (str, optional): 自定义文件名
            timeout (int): 超时时间（秒）

        Returns:
            Optional[list]: 生成图片的路径列表，如果生成失败则返回None
            :param check_stop: 判断是否触发了停止
        """
        try:
            if not prompt_text or not self.is_connected:
                logger.info("警告: 参数无效或未连接")
                return None

            if self.is_generating:
                logger.info("警告: 当前已有任务在运行")
                return None

                # 初始检查
            if check_stop and check_stop():
                return None

            workflow_data, checkpoint = self._load_workflow(task_mode)
            if workflow_data is None:
                return None

            # 准备参数
            if check_stop and check_stop():
                return None

            workflow_data, checkpoint = self._load_workflow(task_mode)
            if workflow_data is None:
                return None

            save_path = Path(output_dir)
            save_path.mkdir(parents=True, exist_ok=True)

            timestamp = time.strftime("%Y%m%d_%H%M%S")
            if custom_filename:
                image_filename = f"{custom_filename}.png"
            else:
                safe_prompt = "".join(x for x in prompt_text[:30] if x.isalnum() or x in (' ', '_', '-'))
                image_filename = f"{timestamp}_{safe_prompt}.png"

            self._save_path = save_path
            self._image_filename = image_filename
            self._generated_paths = None

            params_to_update = {
                "checkpoint_data_json_workflow": model_name if model_name else checkpoint,
                "prompt_data_json_workflow": prompt_text,
                "negative_data_json_workflow": negative_prompt_text if negative_prompt_text else "",
                "width_data_json_workflow": width,
                "height_data_json_workflow": height,
                "sampling_method_data_json_workflow": sampling_method,
                "steps_data_json_workflow": steps,
                "batch_count_data_json_workflow": batch_count,
                "batch_size_data_json_workflow": batch_size,
                "cfg_scale_data_json_workflow": cfg_scale,
                "seed_data_json_workflow": random.randint(0, 2**32 - 1),
            }

            if image_name:
                params_to_update["image_data_json_workflow"] = image_name

            self._update_workflow_params(workflow_data, params_to_update)

            try:
                response = requests.post(
                    f"http://{self.server_url}/prompt",
                    json=workflow_data,
                    timeout=timeout
                )
                if not response.ok:
                    return None

                # 获取并保存prompt_id
                response_data = response.json()
                self.current_prompt_id = response_data.get("prompt_id")
                self.is_generating = True

                generation_successful = threading.Event()

                def save_callback(success: bool):
                    generation_successful.set()

                self._save_callback = save_callback

                # 在等待生成的过程中定期检查状态
                start_time = time.time()
                while not generation_successful.is_set():
                    if check_stop and check_stop():
                        self.stop_current_task()
                        return None

                    if time.time() - start_time > timeout:
                        logger.error("生成任务超时")
                        self.stop_current_task()
                        return None

                    await asyncio.sleep(0.1)  # 每 100ms 检查一次

                return self._generated_paths

            except Exception as e:
                if check_stop and check_stop():
                    return None
                logger.error(f"运行任务失败: {e}")
                return None

            finally:
                # 清理状态和临时属性
                self.is_generating = False
                self.current_prompt_id = None
                if hasattr(self, '_save_callback'):
                    delattr(self, '_save_callback')
                if hasattr(self, '_save_path'):
                    delattr(self, '_save_path')
                if hasattr(self, '_image_filename'):
                    delattr(self, '_image_filename')
                if hasattr(self, '_generated_paths'):
                    delattr(self, '_generated_paths')

        except Exception as e:
            logger.error(f"运行任务失败: {e}")
            self.is_generating = False
            self.current_prompt_id = None
            return None